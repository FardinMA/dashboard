import streamlit as st
import pandas as pd
import numpy as np
import pickle
import io
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


st.set_page_config(page_title="TekhLeads", layout="wide")
st.title("Leads Dashboard")
st.write("Please select your desired options:")

@st.cache_data
def load_data():
    df = pd.read_csv("Lead database CSV.csv", engine="python", encoding="utf-8")
    df.columns = df.columns.str.strip()
    return df

@st.cache_resource
def load_model():
    with open("leadmodel.pkl", "rb") as f:
        return pickle.load(f)

df = load_data()
model = load_model()

def raw2feature(df_raw):
    df = df_raw.copy()

    df["is_private_investor_group_member"] = 0
    df["has_linkedin"] = 0
    df["has_interacted_content"] = 0
    df["has_booking_option"] = 0

    for index in df.index:

        if str(df.loc[index, "Member of Private Investor Group?"]).lower() == "yes":
            df.loc[index, "is_private_investor_group_member"] = 1

        if str(df.loc[index, "Has a LinkedIn?"]).lower() == "yes":
            df.loc[index, "has_linkedin"] = 1

        if str(df.loc[index, "Interacted with Relevant Content?"]).lower() == "yes":
            df.loc[index, "has_interacted_content"] = 1

        if str(df.loc[index, "Option to request service/book appointment?"]).lower() == "yes":
            df.loc[index, "has_booking_option"] = 1

    df["has_email"] = 0
    df["has_phone"] = 0
    df["has_company_website"] = 0
    df["has_personal_website"] = 0
    df["has_other_social_media"] = 0

    for index in df.index:

        if pd.notna(df.loc[index, "Email"]) and str(df.loc[index, "Email"]).strip() != "":
            df.loc[index, "has_email"] = 1

        if pd.notna(df.loc[index, "Phone Number"]) and str(df.loc[index, "Phone Number"]).strip() != "":
            df.loc[index, "has_phone"] = 1

        if pd.notna(df.loc[index, "Company Website"]) and str(df.loc[index, "Company Website"]).strip() != "":
            df.loc[index, "has_company_website"] = 1

        if pd.notna(df.loc[index, "Personal Website"]) and str(df.loc[index, "Personal Website"]).strip() != "":
            df.loc[index, "has_personal_website"] = 1

        if pd.notna(df.loc[index, "Other Social Media"]) and str(df.loc[index, "Other Social Media"]).strip() != "":
            df.loc[index, "has_other_social_media"] = 1

    df["engagement_level_numeric"] = 1

    for index in df.index:
        level = str(df.loc[index, "Engagement Level"]).lower()

        if level == "medium":
            df.loc[index, "engagement_level_numeric"] = 2
        elif level == "high":
            df.loc[index, "engagement_level_numeric"] = 3

    profession_dummies = pd.get_dummies(
        df["Profession"],
        prefix="profession"
    )

    sector_dummies = pd.get_dummies(
        df["Sector"],
        prefix="sector"
    )

    df = pd.concat(
        [df, profession_dummies, sector_dummies],
        axis=1
    )

    drop_cols = ["Name","Sector", "Profession","Company/Network","Company Size","Company Website",
        "Personal Website","Has a LinkedIn?","LinkedIn Profile","Other Social Media","Member of Private Investor Group?",
        "Interacted with Relevant Content?","Option to request service/book appointment?","Engagement Level",
        "Email","Phone Number","Notes","Total","Classification"
    ]

    for col in drop_cols:
        if col in df.columns:
            df = df.drop(columns=col)

    return df

score_cols = [
    "Investor Relevance Score",
    "Contactability Score",
    "Interest Level Score"
]

needs_scoring = df[score_cols].isna().any(axis=1)

if needs_scoring.any():
    features = raw2feature(df.loc[needs_scoring])
    expected_features = model.estimators_[0].get_booster().feature_names
    features = features.reindex(columns=expected_features, fill_value=0)
    preds = model.predict(features)
    df.loc[needs_scoring, score_cols] = np.round(preds).astype(int)


df["Total"] = df[score_cols].mean(axis=1).round(0).astype(int)
df["Classification"] = "RED"
df.loc[df["Total"] > 5, "Classification"] = "GREEN"

col1, col2 = st.columns(2)
with col1:
    sector = st.selectbox("Sector", ["All"] + sorted(df["Sector"].dropna().unique().tolist()))
with col2:
    profession = st.selectbox("Profession", ["All"] + sorted(df["Profession"].dropna().unique().tolist()))

col3, col4 = st.columns(2)
with col3:
    has_linkedin = st.checkbox("Has LinkedIn")
with col4:    
    interacted = st.checkbox("Has interacted with relevant content previously")

private_group = st.checkbox("Member of Company/Network/Private Investor Group?")
comp_size = st.selectbox("If yes, please specify a size:", ["All", "2-10", "11-50", "51-200", "201-500"])

col5, col6 = st.columns(2)
with col5:
    green_list = st.checkbox("Green list")
with col6:
    all_list = st.checkbox("All")

filtered = df.copy()

if not all_list:

    if sector != "All":
        filtered = filtered[filtered["Sector"] == sector]

    if profession != "All":
        filtered = filtered[filtered["Profession"] == profession]

    if comp_size != "All":
        filtered = filtered[filtered["Company Size"] == comp_size]

    if has_linkedin:
        filtered = filtered[filtered["Has a LinkedIn?"].astype(str).str.lower() == "yes"]

    if interacted:
        filtered = filtered[filtered["Interacted with Relevant Content?"].astype(str).str.lower() == "yes"]

    if private_group:
        filtered = filtered[filtered["Member of Private Investor Group?"].astype(str).str.lower() == "yes"]

    if green_list:
        filtered = filtered[filtered["Classification"] == "GREEN"]


st.subheader("Your Filtered Leads")
st.dataframe(filtered, use_container_width=True)

csv_out = filtered.to_csv(index=False).encode("utf-8")

if "Notes" in filtered.columns:
    filtered = filtered.assign(
        _has_notes=filtered["Notes"].notna() & (filtered["Notes"].astype(str).str.strip() != "")
    ).sort_values(
        by="_has_notes",
        ascending=False
    ).drop(columns="_has_notes")

def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")

        workbook = writer.book
        worksheet = writer.sheets["Leads"]

        classification_col_idx = df.columns.get_loc("Classification") + 1
        classification_col_letter = chr(64 + classification_col_idx)

        start_row = 2
        end_row = len(df) + 1
        end_col_letter = chr(64 + len(df.columns))

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_rule = FormulaRule(
            formula=[f"${classification_col_letter}{start_row}=\"GREEN\""],
            fill=green_fill
        )

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_rule = FormulaRule(
            formula=[f"${classification_col_letter}{start_row}=\"RED\""],
            fill=red_fill
        )

        worksheet.conditional_formatting.add(
            f"A{start_row}:{end_col_letter}{end_row}",
            green_rule
        )
        worksheet.conditional_formatting.add(
            f"A{start_row}:{end_col_letter}{end_row}",
            red_rule
        )

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = f"A1:{end_col_letter}{end_row}"

    return output.getvalue()

excel_bytes = dataframe_to_excel_bytes(filtered)

st.download_button(
    label="Download Lead List",
    data=excel_bytes,
    file_name="Your Lead List.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

